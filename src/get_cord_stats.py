"""Collect feature statistics from JSONL files and update Excel spreadsheet."""

import os
import sys
import glob
from copy import copy
from datetime import datetime

from openpyxl import load_workbook

from json2attribute import json2attribute

json_parser = json2attribute("sz_default_config.json")


if len(sys.argv) != 2:
    print("\nsyntax:\n\tpython3 get_cord_stats <directory>\n")
    sys.exit(1)


file_spec = sys.argv[1]
if os.path.isdir(file_spec):
    dir_name = file_spec
    file_spec += os.path.sep + "*.jsonl"
else:
    dir_name = os.path.dirname(file_spec)

file_list = glob.glob(file_spec)
if len(file_list) == 0:
    print("\nno files found!\n")
    sys.exit(1)

stats_file = dir_name + os.path.sep + "_CORD_STATS.xlsx"
if not os.path.exists(stats_file):
    print(f"\n{stats_file} not found!\n")
    sys.exit(1)

wb = load_workbook(stats_file)
ws = wb.worksheets[0]
column_header = list(next(ws.values))

record_types = []
features = []
stat_pack = {}

any_updates = False

for file_name in file_list:
    with open(file_name, "r", encoding="utf-8") as f:
        source, geo = os.path.basename(file_name).replace(".jsonl", "").split("-")
        column_values = {"SOURCE": source, "GEO": geo, "RECORD_COUNT": 0}
        for line in f:
            attr_list = json_parser.parse(line)
            for attr_data in attr_list:
                if not attr_data.get("FTYPE_CODE"):
                    continue
                if attr_data["FTYPE_CODE"] == "RECORD_TYPE":
                    column_name = f"{attr_data['ATTR_VALUE']}_COUNT"
                else:
                    column_name = f"{attr_data['FTYPE_CODE']}_FEATURES"
                if column_name in column_values:
                    column_values[column_name] += 1
                else:
                    column_values[column_name] = 1
            column_values["RECORD_COUNT"] += 1
            if column_values["RECORD_COUNT"] % 100000 == 0:
                print(f"{column_values['RECORD_COUNT']} rows processed for {file_name}")

        print(f"{column_values['RECORD_COUNT']} rows processed for {file_name}, complete!")

        row_found = False
        row_idx = 0
        for row in ws.iter_rows():
            row_idx += 1
            if row[column_header.index("SOURCE")].value == source and row[column_header.index("GEO")].value == geo:
                row_found = True
                print("FOUND!")
                break

        if not row_found:  # insert new row with formatting
            last_row = ws[row_idx]
            row_idx += 1
            ws.insert_rows(idx=row_idx, amount=1)
            row = ws[row_idx]
            for i, last_cell in enumerate(last_row):
                if last_cell.has_style:
                    row[i].font = copy(last_cell.font)
                    row[i].border = copy(last_cell.border)
                    row[i].fill = copy(last_cell.fill)
                    row[i].number_format = copy(last_cell.number_format)
                    row[i].protection = copy(last_cell.protection)
                    row[i].alignment = copy(last_cell.alignment)
                row[i].value = 0
        else:
            row = ws[row_idx]

        updated = False
        for column_name, column_value in column_values.items():
            if column_name not in column_header:
                column_header.append(column_name)
                # ws.insert_cols()
                ws.cell(row=1, column=len(column_header)).value = column_name
                ws.cell(row=row_idx, column=len(column_header)).value = column_value
                # row[column_header.index(column_name)].value = column_value
                updated = True
            elif row[column_header.index(column_name)].value != column_value:
                row[column_header.index(column_name)].value = column_value
                updated = True
        for column_name in column_header:  # remove values no longer present
            if column_name in column_values or column_name == "LAST_UPDATED":
                continue
            if row[column_header.index(column_name)].value:
                row[column_header.index(column_name)].value = 0
                updated = True

        if updated:
            row[column_header.index("LAST_UPDATED")].value = datetime.today()
            print(f"-->> updated {file_name}")
            any_updates = True


if any_updates:
    backup_file = stats_file + ".bak"
    if os.path.exists(backup_file):
        os.remove(backup_file)
    os.rename(stats_file, backup_file)
    wb.save(stats_file)
    print("updates saved!")
